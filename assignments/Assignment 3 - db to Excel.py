#Refered to Rahul Koushik PES1201700121's file

#Please run ex2.py first to create database object and enter values into portfolios before running this


from openpyxl import Workbook

from sqlalchemy import Column,Integer, String, Float,Date, create_engine, exc, orm, MetaData, Table
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker

engine=create_engine("sqlite:///ex2.db") 
engine.connect()
session= sessionmaker(bind=engine)()

meta = MetaData()

meta.create_all(engine)



Base=declarative_base()

Shortlist=Table(
	'Shortlist', meta, 
	Column('Name', String(10), primary_key = True),	
	Column('Percentage', Integer), 
)

class ShortlistExport(Base):
	__tablename__="Shortlist"
	pfname=Column(String(10), primary_key=True)
	stk=Column(String(10), primary_key=True)	
	no=Column(Float)
	def __init__(self,Name,stk,no):
		self.Name= Name
		self.Percentage=Percentage
	def __str__(self):
		return self.Name+"|"+str(self.Percentage)

wb = Workbook()
ws = wb.active

sheets=[]
var = 1

	
result=[r for r in session.query(Shortlist).all()]
for i in result:

	sheets = ws.cell(row=var, column=1)
	sheets.value = i.pfname	

	sheets = ws.cell(row=var, column=2)
	sheets.value = i.Percentage

	var+=1
wb.save('Portfolios.xlsx')

