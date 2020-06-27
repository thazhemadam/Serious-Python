

from tkinter import *
from tkinter.messagebox import showinfo, showerror
from datetime import date
import sqlite3

#References:s
# python_3_email_with_attachment.py
# Created by Robert Dempsey on 12/6/14.
# Copyright (c) 2014 Robert Dempsey. Use at your own peril.



def send_mail():
    '''This is to send a mail'''
    from sys import exc_info
    import smtplib
    from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from os import path, getcwd

    
    sender = 'sender@gmail.com'
    gmail_password = 'password'
    recipients = ['recipient1@gmail.com', 'recipient2@gmail.com']

    # Create the enclosing (outer) message
    outer = MIMEMultipart()
    outer['Subject'] = '[CODE]Mailing Code'
    outer['To'] = COMMASPACE.join(recipients)
    outer['From'] = sender
    outer.preamble = 'You will not see this in a MIME-aware mail reader.\n'

    # List of attachments
    attachments = [getcwd() + '/Shortage.xlsx']

    # Add the attachments to the message
    for file in attachments:
        try:
            with open(file, 'rb') as fp:
                msg = MIMEBase('application', "octet-stream")
                msg.set_payload(fp.read())
            encoders.encode_base64(msg)
            msg.add_header('Content-Disposition', 'attachment', filename=path.basename(file))
            outer.attach(msg)
        except:
            print("Unable to open one of the attachments. Error: ", exc_info()[0])
            raise
        break

    composed = outer.as_string()

    # Send the email
    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as s:
            s.ehlo()
            s.starttls()
            s.ehlo()
            s.login(sender, gmail_password)
            s.sendmail(sender, recipients, composed)
            s.close()
        showinfo("DELIVERY REPORT", "EMAIL SENT!!")
    except:
        showerror("DELIVERY REPORT", "Error: {}".format(sys.exc_info()[0]))



def drop_tables(cur):
    '''Drops any existing table'''
    cur.execute("""drop table if exists Student""")
    cur.execute("""drop table if exists Teacher""")




def create_tables(cur):
    '''Creates tables'''
    cur.execute("Select * from sqlite_master where type = 'table'")
    if cur.fetchone() is None:
        cur.execute("""CREATE table Student 
                    (Name  VARCHAR(25),
                    USID INTEGER PRIMARY KEY,
                    Password VARCHAR(25),
                    Courses VARCHAR(25),
                    Email VARCHAR(25),
                    Attended INTEGER,
                    Total INTEGER)""") #Add email here

        cur.execute("""CREATE table Teacher 
                    (Name  VARCHAR(25),
                    USID INTEGER PRIMARY KEY,
                    Password VARCHAR(25),
                    Courses VARCHAR(25),
                    Email VARCHAR(50))""")

    #cur.execute("Select * from sqlite_master where type = 'table'")
    print("USER.db is Opened...")


    connec.commit()



def insert(ut, data):
    '''Inserting data to the table'''
    if ut == 1:
        cur.execute("""insert into Student VALUES (? , ?, ?, ?, ?, ?, ?)""", data)
    elif ut == 2:
        cur.execute("""insert into Teacher VALUES (?, ?, ?, ?, ?)""", data)
    connec.commit()



def updates(data, check_button):
    att, classes = 0, 0
    for i in range(len(data)):
        att = data[i][1] + check_button[i].get()
        classes = data[i][2] + 1
        cur.execute("""UPDATE Student SET Attended = '{}', Total = '{}' WHERE Name = '{}'""".format(att, classes,data[i][0]))
    connec.commit()
    att_window_f.destroy()
    second_screen_faculty()



def validate(value, item, passwrd=''):
    value = value.split(" ")
    if value[1] == 'USID':
        if value[0] == 'Student':
            cur.execute("SELECT USID from Student where USID = '{}'".format(item))
        elif value[0] == 'Teacher':
            cur.execute("SELECT USID from Teacher where USID = '{}'".format(item))

        usid = cur.fetchall()
        if len(usid) >= 1:
            return True

    elif value[1] == 'Password':
        if value[0] == 'Student':
            cur.execute("SELECT Password from Student where USID = '{}'".format(item))
        elif value[0] == 'Teacher':
            cur.execute("SELECT Password from Teacher where USID = '{}'".format(item))

        checker = cur.fetchall()[0][0]
        if passwrd == checker:
            return True
    else: return False



def shortage():
    cur.execute("SELECT Name, USID, Courses, Attended, Total FROM Student")
    data = cur.fetchall()
    file = "Shortage.xlsx"
    from openpyxl import Workbook
    wb = Workbook()
    print("Workbook Created...")
    ws = wb.create_sheet("Attendance Shortage", 0)
    cell_data = ["Name", "ID", "Course", "Attended", "Total Classes", "Percentage"]
    for i in range(len(cell_data)):
        ws[str(chr(65 + i)) + '1'] = cell_data[i]

    for row in range(len(data)):
        percentage = (data[row][3]/data[row][4])*100
        print(percentage)
        if percentage < 85:
            for col in range(5):
                ws[str(chr(65 + col)) + str(row + 2)] = data[row][col]
        ws[str(chr(70)) + str(row + 2)] = percentage

    wb.save(file)
    send_mail()



def login(window):
    global usn
    usn = username.get()
    pswd = password.get()
    usid_call = lambda x: validate(x, int(usn))
    pass_call = lambda  x: validate(x, int(usn), pswd)

    if usid_call("Student USID") and pass_call("Student Password"):
        window.destroy()
        second_screen_student()

    elif usid_call("Teacher USID") and pass_call("Teacher Password"):
        window.destroy()
        second_screen_faculty()

    else:
        showerror("LOGIN TRY", "INVALID USID/PASSWORD")
        main_window.destroy()
        main_screen_window()



def submit(data):
    insert(1, data)
    return login(register_window)



def register():
    main_window.destroy()
    global register_window
    register_window = Tk()
    register_window.title("Attendance Report")
    register_window.geometry("300x350")
    Label(register_window).pack(fill="both", expand=True)
    Label(text = "Enter your details below", font = ("ARIAL", 15)).pack()
    Label(register_window).pack()
    global name, username, password, email
    name, username, password, email, course = StringVar(), IntVar(), StringVar(), StringVar(), StringVar()
    Label(text = "Name: ").pack()
    Entry(register_window, textvariable = name).pack()
    Label(text="USID: ").pack()
    Entry(register_window, textvariable = username).pack()
    Label(text="Password: ").pack()
    Entry(register_window, textvariable = password).pack()
    Label(text="Email: ").pack()
    Entry(register_window, textvariable = email).pack()
    Label(text="Course: ").pack()
    Entry(register_window, textvariable=course).pack()
    Label(register_window).pack()
    submit_data = lambda : submit([name.get(), username.get(), password.get(), course.get(), email.get(), 0, 0])
    Button(text = "Submit", height = "1", width = "15", activebackground = "yellow", bg = "orange", command = submit_data).pack()
    Label(register_window).pack()
    register_window.mainloop()
    connec.commit()



def second_screen_student():
    global second_window
    second_window = Tk()
    second_window.title("Student Home Page")
    second_window.geometry("300x300")

    main_lab = Label(text = "USN = {}".format(usn), bg = "orange", font = ("ARIAL", 20))
    main_lab.config(anchor = CENTER)
    main_lab.pack()
    Label(second_window).pack()
    Button(text="Registered Courses", height = "2", width = "15", activebackground = "yellow", bg = "orange", command = courses_screen_student).pack()
    Label(second_window).pack()
    Button(text="Attendance", height="2", width="15", activebackground = "yellow", bg = "orange", command = attendance_screen_student).pack()
    Label(second_window).pack()
    global search
    search = StringVar()
    Entry(second_window, textvariable = search).pack()
    Button(text="Search", height="1", width="15", activebackground = "yellow", bg = "orange", command = search_screen).pack()
    Label(second_window).pack()
    Button(text="Sign Out", height="1", width="15", activebackground = "yellow", bg = "orange", command = second_window.destroy).pack()
    Label(second_window).pack()
    second_window.mainloop()
    main_screen_window()



def courses_screen_student():
    second_window.destroy()
    global courses_window
    courses_window = Tk()
    courses_window.title("Registered Courses")
    #courses_window.geometry("250x250")
    Label(courses_window).pack(fill = "both", expand = True)

    cur.execute("SELECT Courses from Student WHERE USID = '{}'".format(usn))
    data = cur.fetchall()[0][0]
    main_lab1 = Label(text="Course Name: {}".format(data), bg = "orange", font=("ARIAL", 15))
    main_lab1.config(anchor=CENTER)
    main_lab1.pack()
    Label(courses_window).pack()
    main_lab2 = Label(text="Course Duration: {} upto 5 days".format(date.today()), font=("ARIAL", 10))
    main_lab2.config(anchor=CENTER)
    main_lab2.pack()

    Label(courses_window).pack()
    Button(text="Back", height="2", width="15", activebackground = "yellow", bg = "orange", command = courses_window.destroy).pack()
    Label(courses_window).pack()
    courses_window.mainloop()
    second_screen_student()



def attendance_screen_student():
    second_window.destroy()
    global att_window
    att_window = Tk()
    att_window.title("Attendance Report")
    #att_window.geometry("250x250")
    Label(att_window).pack(fill = "both", expand = True)

    cur.execute("SELECT Courses, Attended, Total from Student WHERE USID = '{}'".format(usn))
    data = cur.fetchall()[0]
    #print(data)
    main_lab1 = Label(text="Course Name: {}".format(data[0]), bg = "orange", font=("ARIAL", 15))
    main_lab1.config(anchor=CENTER)
    main_lab1.pack()
    Label(att_window).pack()
    attendance = 0
    if data[2] is not 0:
        attendance = (data[1] / data[2]) * 100
    main_lab2 = Label(text="Attendance Percentage: {num:.2f}".format(num = attendance), font=("ARIAL", 10))
    main_lab2.config(anchor=CENTER)
    main_lab2.pack()
    Label(att_window).pack()
    Button(text="Back", height="2", width="15", activebackground = "yellow", bg = "orange",command=att_window.destroy).pack()
    Label(att_window).pack()
    att_window.mainloop()
    second_screen_student()



def search_screen():

    data = search.get()
    if data is '':
        showinfo("Search Failed", "No Input Given.")
        return
    cur.execute("SELECT Name from Student where Courses = '{}'".format(data))
    name_list = cur.fetchall()

    if data is not '' and len(name_list):

        global search_window
        search_window = Tk()
        search_window.title("Search Results")
        #search_window.geometry("200x250")
        Label(search_window).pack(fill = "both", expand = True)

        main_lab1 = Label(search_window ,text="Course: {}".format(data), bg = "orange", font=("ARIAL", 15))
        main_lab1.config(anchor=CENTER)
        main_lab1.pack()
        Label(search_window).pack()
        for names in name_list:
            main_lab2 = Label(search_window, text="Student Name: {}".format(names[0]), font=("ARIAL", 10))
            main_lab2.config(anchor=CENTER)
            main_lab2.pack()
            Label(search_window).pack()
        Button(search_window, text = "Back", height = "1", width = "15", activebackground = "yellow", bg = "orange", command = search_window.destroy).pack()
        Label(search_window).pack()
        search_window.mainloop()
    else:
        showinfo("Search Failed", "No Students enrolled in this Course")



def second_screen_faculty():

    global second_window_f
    second_window_f = Tk()
    second_window_f.title("Faculty Home Page")
    second_window_f.geometry("250x300")
    main_lab = Label(text = "USN = {}".format(usn), bg = "orange",font = ("ARIAL", 20))
    main_lab.config(anchor = CENTER)
    main_lab.pack(fill = "both", expand = True)
    Label(second_window_f).pack()
    Button(text="Courses Undertaken", height = "2", width = "15", activebackground = "yellow", bg = "orange", command = courses_screen_faculty).pack()
    Label(second_window_f).pack()
    Button(text="Attendance", height="2", width="15", activebackground = "yellow", bg = "orange", command = attendance_screen_faculty).pack()
    Label(second_window_f).pack()

    global  search
    search = StringVar()
    Entry(second_window_f, textvariable = search).pack()
    Button(text="Search", height="1", width="15", activebackground = "yellow", bg = "orange", command = search_screen).pack()
    Label(second_window_f).pack()
    Button(text="Sign Out", height="1", width="15", activebackground = "yellow", bg = "orange", command=second_window_f.destroy).pack()
    Label(second_window_f).pack()
    second_window_f.mainloop()
    main_screen_window()



def attendance_screen_faculty():
    second_window_f.destroy()
    global att_window_f
    att_window_f = Tk()
    att_window_f.title("Attendance Report")
    #att_window_f.geometry("300x400")
    Label(att_window_f).pack(fill = "both", expand = True)

    cur.execute("SELECT Courses from Teacher WHERE USID = '{}'".format(usn))
    data = cur.fetchall()[0][0]
    main_lab1 = Label(text="Course Name: {}".format(data), bg = "orange", font=("ARIAL", 15))
    main_lab1.config(anchor=CENTER)
    main_lab1.pack()
    Label(att_window_f).pack()

    cur.execute("SELECT Name, Attended, Total from Student WHERE Courses = '{}'".format(data))
    data = cur.fetchall()

    global check_box
    check_box = []

    for i in range(len(data)):
        check_box.append(IntVar())
        Checkbutton(att_window_f, text = "{}".format(data[i][0]), variable = check_box[i], onvalue = 1, offvalue = 0).pack()

    Label(att_window_f).pack()
    Button(text="Export and Send", height="2", width="15", activebackground = "yellow", bg = "orange", command=shortage).pack()
    Label(att_window_f).pack()

    submit_attendance = lambda : updates(data, check_box)
    Button(text="Submit", height="2", width="15", activebackground = "yellow", bg = "orange", command=submit_attendance).pack()
    Label(att_window_f).pack()
    Button(text="Back", height="2", width="15", activebackground = "yellow", bg = "orange", command=att_window_f.destroy).pack()
    Label(att_window_f).pack(fill = "both", expand = True)
    att_window_f.mainloop()
    second_screen_faculty()



def courses_screen_faculty():
    second_window_f.destroy()
    global courses_window_f
    courses_window_f = Tk()
    courses_window_f.title("Registered Courses")
    #courses_window_f.geometry("300x300")
    Label(courses_window_f).pack(fill = "both", expand = True)

    cur.execute("SELECT Courses from Teacher WHERE USID = '{}'".format(usn))
    data = cur.fetchall()[0][0]
    main_lab1 = Label(text="Course Name: {}".format(data), bg = "orange", font=("ARIAL", 15))
    main_lab1.config(anchor=CENTER)
    main_lab1.pack()
    Label(courses_window_f).pack()
    main_lab2 = Label(text="Course Duration: {} until 5 days".format(date.today()), font=("ARIAL", 10))
    main_lab2.config(anchor=CENTER)
    main_lab2.pack()
    Label(courses_window_f).pack()
    Button(text="Back", height="2", width="15", activebackground = "yellow", bg = "orange", command = courses_window_f.destroy).pack()
    Label(courses_window_f).pack()
    courses_window_f.mainloop()
    second_screen_faculty()



def main_screen_window():
    global main_window
    main_window = Tk()
    main_window.title("https://www.pes.edu/")
    main_window.geometry("300x350")
    #main_window.config(background = "white")
    main_lab = Label(text = "PES UNIVERSITY", bg = "orange", font = ("ARIAL BLACK", 20))
    main_lab.config(anchor = CENTER)
    main_lab.pack(fill = "both", expand = True)


    global username, password
    username = StringVar()
    password = StringVar()

    Label(main_window).pack()
    Label(main_window, text = "User Name: ").pack()
    Entry(main_window, textvariable = username).pack()

    Label(main_window).pack()
    Label(main_window, text="Password: ").pack()
    Entry(main_window, textvariable = password, show = "*").pack()

    Label(main_window).pack()
    login_user = lambda : login(main_window)
    Button(text="Login", height = "1", width = "10", activebackground = "yellow", bg = "orange",command = login_user).pack()
    Label(main_window).pack()
    Button(text="Register", height = "1", width = "10", activebackground = "yellow", bg = "orange", command = register).pack()
    Label(main_window).pack()
    Button(text="Exit", height="1", width="10", activebackground = "yellow", bg = "orange", command = exit).pack()
    Label(main_window).pack()
    main_window.mainloop()



if __name__ == "__main__":
    
    connec = sqlite3.connect("USER.db")
    cur = connec.cursor()

    '''Test data added initially'''
    '''
    name1 = [('A', 20181, "A1", "Computer", "None", 0, 0),
             ('B', 20182, "B1", "Computer", "None", 0, 0),
             ('C', 20183, "C1", "Computer", "None", 0, 0)]
    name2 = [('TA', 1, "TA1", "Computer", "recipient1@gmail.com"),
             ('TB', 2, "TB1", "Electronics", "recipient2@hotmail.com"),
             ('TC', 3, "TC1", "Accounts", "recipient3@gmail.com")]
    '''

    create_tables(cur)
    cur.execute("SELECT Name, Attended from Student WHERE Courses = '{}'".format("Computer"))
    print(cur.fetchall())
    COMMASPACE = ''
    '''for i in range(3):
        insert(1, name1[i])
        insert(2, name2[i])
    #For initial addition of data
    '''
    print("\nSTUDENTS:\nName | ID | Pswd |\t Courses| Attended")
    print_ID_PASS = lambda x: tuple(map(lambda i: print(i), x))

    cur.execute("SELECT * from Student")
    data = cur.fetchall()
    print_ID_PASS(data)
    cur.execute("SELECT * from Teacher")
    data = cur.fetchall()
    print("\nTEACHERS:\nName | ID | Pswd |\t Courses| \t\t\t Email|")
    print_ID_PASS(data)
    connec.commit()


    main_screen_window()